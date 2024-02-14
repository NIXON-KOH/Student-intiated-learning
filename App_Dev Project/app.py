from flask import Flask, render_template, request, redirect,url_for
from user import Customer, Admin
from product import Product, Rewards
import shelve, json, random, datetime, os
from Transaction import Transaction
import re
import openpyxl
from flask import flash

def initate():
    global Products,reward, transaction
    reward = []
    Products = []
    transaction = []
    with shelve.open("Databases/product") as f: 
        for i in f:
            #id_, name,Image, Stocks, Desc, Type, Status,categories,RetailCost,gender,ListCost,reviews
            #Reads Product data
            if f[i]["type"] == "product":
                globals()[f[i]['name']] = Product(
                    i,
                    f[i]["name"],
                    f[i]["image"],
                    f[i]["stocks"],
                    f[i]["desc"],
                    f[i]["points"],
                    f[i]["type"],
                    f[i]["categories"],
                    f[i]["RetailCost"],
                    f[i]["gender"],
                    f[i]["ListCost"],
                )
                Products.append(f[i]['name'])
            elif f[i]["type"] == "rewards":
                globals()[f[i]["name"]] = Rewards(
                    i,
                    f[i]["name"],
                    f[i]["image"],
                    f[i]["stock"],
                    f[i]["desc"],
                    f[i]["type"],
                    f[i]["Status"],
                    f[i]["ListCost"]
                ) 
                reward.append(f[i]['name'])
    with shelve.open("Databases/transaction") as f:
        for i in f:
            globals()[i] = Transaction(
                f[i]["Date"],
                f[i]["Product"],
                f[i]["Status"],
                f[i]["Cust"],
                f[i]["Price"]
            )

def chatbot(query):
    lol = "Sorry i could not get that."
    intent = "N/A"
    with open("chatbot.json") as f:
        data = json.load(f)
        f.close()
    for i in range(len(data)):
        for n in data[i]["patterns"]:
            if query.lower() in n.lower():
                return random.choice(data[i]["responses"]),data[i]["tag"]
    return lol,intent
                                    

logged_in = False
app = Flask(__name__)

@app.route("/",methods=["GET","POST"])
def lobby():
    global logged_in
    with shelve.open("Databases/user") as f:
        if logged_in == False:
            username = "NOES"
        else:
            username = user.get_name()
        return render_template("landing.html",username = username)

@app.route("/login",methods=["GET","POST"])
def authenticate():
    global logged_in
    if request.method == "POST":
        username = request.form['name']
        password = request.form['password']
        global user
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]['username'] == username and f[i]["password"] == password:
                    if f[i]["role"] == "admin":
                        user = Admin(i,f[i]["role"],f[i]["username"],f[i]["JoinDate"],f[i]["password"],f[i]["image"],f[i]["email"])
                        logged_in = True    
                        return redirect(url_for('admin_home'))
                    elif f[i]["role"] == "user":
                        user = Customer(i,"user",f[i]["username"],f[i]["JoinDate"],f[i]["password"],f[i]["image"],f[i]["email"],f[i]["address"],f[i]["cart"],f[i]["point"])
                        logged_in = True
                        return redirect(url_for('lobby'))
                    else:
                        return render_template("error.html")
            return render_template("error.html")

    return render_template("login.html")

@app.route("/user")
def user_data():    
    #Read User data
    return render_template("", uname=user.get_name(),pwd=user.get_Password(),image=user.get_image(),email=user.get_Email(),addr=user.get_Address(),cart=user.get_cart(),hist=user.get_History(),joindate=user.get_JoinDate(),points=user.get_Points())

@app.route("/registration",methods=["GET",'POST'])
def registration():
    if request.method == "POST":
        name = request.form['name']
        password = request.form['password']
        email = request.form['email']
        address = request.form['address']

        with shelve.open('Databases/user') as f: 
            x = 1
            for i in f:
                if f[i]['username'] == name :
                    return render_template("keyerror.html")
                elif f[i]['email'] == email :
                    return render_template("keyerror.html")
                x += 1
            #Create User Data
            f[str(x)] = {
                "username" : name,
                "password" : password,
                "role" : "user",
                "image" : f"{str(x)}.jpg",
                "email" : email,
                "address" : address,
                "Paynow" : {"name":"No registered name","number":"No registered Account Number","cvv":"No registered cvv"},
                "Paypal" : {"email":"No Registered email","password":"No Registered password"},
                "JoinDate" : datetime.datetime.today().strftime('%d-%m-%Y'),
                "point" : 0,
                "cart" : []
                }
            


        return redirect(url_for("confirmation"))
    return render_template("register.html")     

@app.route("/confirmation")
def confirmation():
    return render_template("confirmation.html")

@app.route("/cart",methods=["GET",'POST'])
def cart():
    if request.method == "POST":
            try: 
                if request.form["submit"] == "checkout":
                    with shelve.open("Databases/user") as f:
                        for i in f:
                            if f[i]["username"] == user.get_name():
                                points = int(f[i]["point"])
                                username = f[i]["username"]
                                hey = ", ".join(i for i in f[i]["cart"])
                                with shelve.open("Databases/product") as g:
                                    for ii in f[i]["cart"]:
                                        ii = ii.replace("_"," ")
                                        print(ii)
                                        for iii in g:
                                            if ii == g[iii]["name"]:
                                                with shelve.open("Databases/transaction") as h:
                                                    x = 0
                                                    for _id in h:
                                                        x = int(_id) + 1
                                                    print(x)
                                                    h[str(x)] = {
                                                        "id" : x,
                                                        "Date" : str(datetime.date.today()),
                                                        "Product" : ii.replace(' ',"_"),
                                                        "Cust" : username,
                                                        "Status" : "active",
                                                        "Price" : g[iii]["ListCost"]
                                                    }
                                                    try:
                                                        points += g[iii]["points"]
                                                    except:
                                                        print(g[iii]["name"])
                                                
                                                break
                                f[i] = {
                                        "username":f[i]["username"],
                                        "password": f[i]["password"],
                                        "role":"user",
                                        "image": f[i]["image"],
                                        "email": f[i]["email"],
                                        "address": f[i]["address"],
                                        "History": {},
                                        "Paynow" : f[i]["Paynow"],
                                        "Paypal" : f[i]["Paypal"],
                                        "JoinDate":f[i]["JoinDate"],
                                        "point":points,
                                        "cart":[]
                                    }
                            
                                    
                                
                                        
                                                    
                return render_template("order_confirmation.html",data=hey, username=username)
            except:
                with shelve.open("Databases/user") as f:
                    
                    for i in f:
                        if f[i]["username"] == user.get_name():
                            item  = request.form["delete"].replace(" ","_")
                            new_cart = list(f[i]["cart"])
                            new_cart.remove(item)

                            f[i] = {
                                "username":f[i]["username"],
                                "password":f[i]["password"],
                                "role":"user",
                                "image":f[i]["image"],
                                "email":f[i]["email"],
                                "address":f[i]["address"],
                                "Paynow" : f[i]["Paynow"],
                                "Paypal" : f[i]["Paypal"],
                                "JoinDate":f[i]["JoinDate"],
                                "point":f[i]["point"],
                                "cart": new_cart
                                }
                            return redirect(url_for("cart"))
        

    try:         
        with shelve.open("Databases/user") as f:
            for i in f:
                    if user.get_name() == f[i]["username"]:
                        cards = ""
                        items = 0
                        cost = 0.0
                        
                        for ii in f[i]["cart"]:
                            ii = ii.replace("_"," ")
                            with shelve.open("Databases/product") as g:
                                for iii in g:
                                    if g[iii]["name"] == ii.replace("_"," "):
                        
                                        card = f'''<div class="cart-item">
                                    <hr id="cart-item-hr">
                                    <div class="cart-item-left">
                                        <img class="cart-img" src="../static/Assets/{ g[iii]["image"].replace(" ","_") }">
                                    </div>
                                    <div class="cart-item-middle">
                                        <h2>{ g[iii]["name"] }</h2>
                    
                                        <p>Size: S </p>
                                        <span><h2> Cost </h2><p>{ g[iii]["ListCost"]}</p></span>
                                        <h3>Quantity : 1</h3>
                                        <form method="post">
                                        <button id="delete" value="{ g[iii]["name"] }" name="delete">Remove</button>
                                        </form>
                                    </div>
                                </div>'''
                                        cards += card
                                        items += 1
                                        cost += float(g[iii]["ListCost"])
    
        return render_template("Cartpage.html",data=cards,total_items=items, total_price = cost,username = user.get_name())
    except: return render_template("keyerror.html")

@app.route("/admin/<string:Account>",methods=["POST","GET"])
def admin(Account):
    if request.method=="POST":
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == Account:
                    if request.form["username"] != "":
                        username = request.form["username"]
                        image = request.form["username"].replace(" ","_") + ".jpg"
                        os.rename(f"/static/Assets/{f[i]['image']}",f"static/Assets/{image}")
                    else:
                        username = f[i]["username"]
                        image = f[i]["image"]
                    
                    if request.form["email"] != "":
                        email = request.form["email"]
                    else:
                        email = f[i]["email"]
                    
                    if request.form["password"] != "":
                        password = request.form["password"]
                    else:
                        password = f[i]["password"]
                    uploaded_file = request.files["file"]
                    
                    if uploaded_file.filename != "":
                        os.remove(f"static/Assets/{f[i][image]}")
                        uploaded_file.save(f"static/Assets/{image}")



                    f[i] = {
                        "username":username,
                        "password":password,
                        "JoinDate":f[i]["JoinDate"],
                        "image" : image,
                        "role":"admin",
                        "email":email
                        }
                    
    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == Account:
                username = f[i]["username"]
                image = f[i]["image"]
                JoinDate = f[i]["JoinDate"]
                email = f[i]["email"]
                password = f[i]["password"]

    return render_template("admin_profile.html",username=username,image=image,JoinDate=JoinDate,email=email,password=password)

        

@app.route("/shop",methods=["POST","GET"])
def shop():
    if request.method == "POST":
        #Button will all have same name but different values
        #adds items to cart
        try:
            if logged_in == True:
                with shelve.open("Databases/user") as f:
                    
                    for i in f:
                        if f[i]["username"] == user.get_name():
                            item  = request.form["item"].replace(" ","_")
                            new_cart = list(f[i]["cart"])
                            new_cart.append(item)

                            f[i] = {
                                "username":f[i]["username"],
                                "password":f[i]["password"],
                                "role":"user",
                                "image":f[i]["image"],
                                "email":f[i]["email"],
                                "address":f[i]["address"],
                                "Paynow":f[i]["Paynow"],
                                "Paypal":f[i]["Paypal"],
                                "JoinDate":f[i]["JoinDate"],
                                "point":f[i]["point"],
                                "cart": new_cart
                                }
            else: return render_template("keyerror.html")
        except:
            return render_template("keyerror.html")

                   
    
    cards = ""
    for i in Products:
        card_part_1 = f'''<div class="grid-item">
            <form method="post">
                <div class="product-card">
                    <img class="product-image" src="/static/Assets/{globals()[i].get_Image().replace(" ","_")}" alt="{globals()[i].get_Image()}">
                    <div class="product-details">
                      <div class="product-name"><a href="/item/{globals()[i].get_name()}">{globals()[i].get_name()}</a></div>
                      <span class="icon-cart">  
                          <button name='item' value="{globals()[i].get_name()}"><i class="fa fa-shopping-cart fa-2x" id="icon-cart"></i></button>
                      </span>
                    </div>
                    <div class="product-size">
                      <label for="size">Size:</label>
                      <select id="size" name="size"></form>'''
        sizings = ["S","M","L","XL","XXL"]
        card_part_2 = ""
        for n in sizings:
            if (globals()[i].get_Stocks()).get(n) != 0:
                card_part_2 += f"<option value='{n}'>{n}</option>"
            else:
                card_part_2 += f"<option value='{n}' disabled>{n}</option>"

        card_part_3 = f'''</select>
                    </div>
                    <div class="product-price">Price: {globals()[i].get_ListCost()}</div>
                </div>
            </form>
        </div>'''
        card = card_part_1 + card_part_2 + card_part_3
        cards += card
    try:     
        return render_template("shop.html",cards=cards,username=user.get_name())   
    except:
        return render_template("shop.html",cards=cards)
    

@app.route("/reward",methods=["GET","POST"])
def reward():
    if request.method == "POST":
        try:
            with shelve.open("Databases/user") as f:
                for i in f:
                    if f[i]["username"] == user.get_name():
                        print("Hello")
                        item  = request.form["redeem"].replace(" ","_")
                        print("hello1")
                        new_cart = list(f[i]["cart"])
                        print("Hello2")
                        new_cart.append(item)
                        print("Hellos")
                        f[i] = {
                            "username":f[i]["username"],
                            "password":f[i]["password"],
                            "role":"user",
                            "image":f[i]["image"],
                            "email":f[i]["email"],
                            "address":f[i]["address"],
                            "Paynow": f[i]["Paynow"],
                            "Paypal": f[i]["Paypal"],
                            "JoinDate":f[i]["JoinDate"],
                            "point":f[i]["point"],
                            "cart":new_cart
                            }
                        
                        
        except:
            return render_template("keyerror.html")
    with shelve.open("Databases/product") as f:
        data = ""
        for i in f:
            if f[i]["type"] == "reward":
                card = f'''<div class="reward">
                <img src="/static/Assets/{f[i]["image"]}" alt="{f[i]["name"]}">
                <div class="Info">
                    <header>{f[i]["name"]}</header>
                    <p>Description:
                        <br>
                        {f[i]["Desc"]}
                    </p>
                    <p>{f[i]["ListCost"]}</p>
                    <form method="post">
                    <button type="submit" name="redeem" value="{f[i]["name"]}">Redeem</button></form>
                    </div>
                    </div>'''
                data += card
    try:
        return render_template("reward.html",data=data,username=user.get_name())
    except:
        return render_template("reward.html",data=data)
    

@app.route("/item/<string:item>",methods=["POST","GET"])
def items(item):
    if request.method == "POST":
        
            if logged_in == True:
                with shelve.open("Databases/user") as f:
                    for i in f:
                        if f[i]["username"] == user.get_name():
                            item  = item.replace(" ","_")
                            new_cart = list(f[i]["cart"])
                            new_cart.append(item)

                            f[i] = {
                                "username":f[i]["username"],
                                "password":f[i]["password"],
                                "role":"user",
                                "image":f[i]["image"],
                                "email":f[i]["email"],
                                "address":f[i]["address"],
                                "Paynow" : f[i]["Paynow"],
                                "Paypal" : f[i]["Paypal"],
                                "JoinDate":f[i]["JoinDate"],
                                "point":f[i]["point"],
                                "cart": new_cart
                                }
                            return render_template(url_for("cart"))
            else: 
            
                return render_template("keyerror.html")
    

    with shelve.open("Databases/product") as f:
        for i in f:
            if f[i]["name"] == item:
                name = f[i]["name"]
                image = f[i]["image"]
                stocks = f[i]["stocks"]
                desc = f[i]["desc"]
                points =f[i]["points"]
                categories = f[i]["categories"]
                price = f[i]["ListCost"]
    try:
        return render_template("product_page.html",username=user.get_name(),name=name,image=image.replace(" ","_"),stocks=stocks,points=points,desc=desc,price=price,categories=categories)
    except:
        return render_template("product_page.html",name=name,image=image.replace(" ","_"),stocks=stocks,points=points,desc=desc,price=price,categories=categories)



@app.route("/admin/create-admin",methods=["GET","POST"])
def create_admin():
    if request.method == "POST":
        name = request.form["username"]
        date = str(datetime.datetime.now().strftime("%d-%m-%Y"))
        role = "admin"
        image = request.form["username"]+".jpg"
        password = request.form["password"]
        email = request.form["email"]

        with shelve.open("Databases/user") as f:
            x = 0
            for i in f:
                x = i
            x += 1
            f[str(x)] = {
                        "username" : name,
                        "password": password,
                        "JoinDate":date,
                        "image":image,
                        "role":"admin",
                        "email":email
                }
            
    return render_template("create-admin.html")


@app.route("/admin-create-product",methods=["GET","POST"])
def create_product():
    #Creating product
    if request.method =="POST":
        with shelve.open("Databases/product") as f:
            for i in f:
                n = i
            f[str(n)] = {
                    n,
                    request.form["name"],
                    request.form["image"],
                    request.form["stocks"],
                    request.form["desc"],
                    request.form["points"],
                    request.form["type"],
                    request.form["categories"],
                    request.form["RetailCost"],
                    request.form["gender"],
                    request.form["ListCost"],
                    request.form["reviews"]
            }
            
        
            globals()[request.form["name"]] = Product(
                    n,
                    request.form["name"],
                    request.form["image"],
                    request.form["stocks"],
                    request.form["desc"],
                    request.form["points"],
                    request.form["type"],
                    request.form["categories"],
                    request.form["RetailCost"],
                    request.form["gender"],
                    request.form["ListCost"],
                    request.form["reviews"]
            )
            Products.append(request.form["name"])
            return render_template("admin-rewards.html")
    return render_template("admin_product_create.html")

@app.route("/admin-rewards",methods=["POST","GET"])
def admin_rewards():
    cards = ''
    for i in Rewards:
        id_ = globals()[i].get_id_()
        name = globals()[i].get_name()
        image = globals()[i].get_image()
        stocks = globals()[i].get_Stocks()
        desc = globals()[i].get_desc()
        typed = globals()[i].get_Type()
        points = globals()[i].get_points()
        status = globals()[i].get_status()
        card = name + image + stocks + desc + typed +points + status
        cards += f"{id_} : " + card

    return render_template("admin_reward.html")

@app.route("/admin-rewards/<reward>",methods=["POST","GET"])
def admin_reward_data(reward):
    if request.method == "POST":
        with shelve.open("Databases/product") as f:
            for i in f:
                if f[i]["name"] == reward:
                    if request.form["name"] != "":    
                        name = request.form["name"]
                    else:
                        name = f[i]["name"]
                    
                    if request.form["price"] != "":    
                        price = request.form["price"]
                    else:
                        price = f[i]["ListCost"]
                    
                    if request.form["stock"] != "":    
                        stock = request.form["stock"]
                    else:
                        stock = f[i]["stock"]
                    if request.form["status"] != "":    
                        status = request.form["status"]
                    else:
                        status = f[i]["Status"]
                    
                    if request.form["desc"] != "":    
                        desc = request.form["desc"]
                    else:
                        desc = f[i]["Desc"]
                    
                    f[i] = {
                        "name" : name,
                        "image" : f"{name.replace(' ','_')}.jpg",
                        "stock" : stock,
                        "Desc" : desc,
                        "type" : "reward",
                        "ListCost" : price,
                        "Status" : status 
                    }
                        
                    
                
    with shelve.open("Databases/product") as f:
        for i in f:
            if f[i]["name"] == reward:
                points = f[i]["ListCost"]
                stock = f[i]["stock"]
                status = f[i]["Status"]
                desc = f[i]["Desc"]
    return render_template("admin_reward_edit.html",name=reward,stock=stock,status=status,points=points,desc=desc)

@app.route("/admin/products",methods=["POST","GET"])
def admin_products():
    if request.method == "POST":
        #Deletes Product from the product database
        with shelve.open("Databases/pq  roduct") as f:
            #All button will have name="rm_product" and a value that corrosponds
            for i in f:
                try:
                    if f[i]["name"] == request.form[f[i]["name"]]:
                        del f[request.form["rm_product"]] 
                except:
                    pass
    #Admin can view all the products from the product database
    cards = ""
    with shelve.open("Databases/product") as f:
        for i in f:
            if f[i]["type"] == "product":
                card = f'''
                <tr>
                    <td> {f[i]["name"]} </td>
                    <td> {f[i]["gender"]} </td>
                    <td> {f[i]["categories"]} </td>
                    <td> {f[i]["color"]} </td>
                    <td> {f[i]["stocks"]["S"]} </td>
                    <td> {f[i]["stocks"]["M"]} </td>
                    <td> {f[i]["stocks"]["L"]} </td>
                    <td> {f[i]["stocks"]["XL"]} </td>
                    <td> {f[i]["stocks"]["XXL"]} </td>
                    <td> {f[i]["stocks"]["S"] + f[i]["stocks"]["M"] + f[i]["stocks"]["L"] + f[i]["stocks"]["XL"] + f[i]["stocks"]["XXL"]} </td>
                    <td> {f[i]["ListCost"]} </td>
                    <td><a href="/admin-edit-info/{f[i]["name"]}"><svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="currentColor" class="bi bi-pencil-square" viewBox="0 0 16 16">
  <path d="M15.502 1.94a.5.5 0 0 1 0 .706L14.459 3.69l-2-2L13.502.646a.5.5 0 0 1 .707 0l1.293 1.293zm-1.75 2.456-2-2L4.939 9.21a.5.5 0 0 0-.121.196l-.805 2.414a.25.25 0 0 0 .316.316l2.414-.805a.5.5 0 0 0 .196-.12l6.813-6.814z"/>
  <path fill-rule="evenodd" d="M1 13.5A1.5 1.5 0 0 0 2.5 15h11a1.5 1.5 0 0 0 1.5-1.5v-6a.5.5 0 0 0-1 0v6a.5.5 0 0 1-.5.5h-11a.5.5 0 0 1-.5-.5v-11a.5.5 0 0 1 .5-.5H9a.5.5 0 0 0 0-1H2.5A1.5 1.5 0 0 0 1 2.5z"/>
</svg></a></td>

                    <form method="post">
    <td>
        <button type="submit" class="tgt" name="{f[i]['name']}">
            <svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="currentColor" class="bi bi-trash" viewBox="0 0 16 16">
                <path d="M5.5 5.5A.5.5 0 0 1 6 6v6a.5.5 0 0 1-1 0V6a.5.5 0 0 1 .5-.5m2.5 0a.5.5 0 0 1 .5.5v6a.5.5 0 0 1-1 0V6a.5.5 0 0 1 .5-.5m3 .5a.5.5 0 0 0-1 0v6a.5.5 0 0 0 1 0z"/>
                <path d="M14.5 3a1 1 0 0 1-1 1H13v9a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V4h-.5a1 1 0 0 1-1-1V2a1 1 0 0 1 1-1H6a1 1 0 0 1 1-1h2a1 1 0 0 1 1 1h3.5a1 1 0 0 1 1 1zM4.118 4 4 4.059V13a1 1 0 0 0 1 1h6a1 1 0 0 0 1-1V4.059L11.882 4zM2.5 3h11V2h-11z"/>
            </svg>
        </button>
    </td>
</form>


                </tr>'''
                cards += card
    return render_template("admin_product.html", cards=cards)
        
@app.route("/admin-edit-info/<string:product>", methods=["POST","GET"])
def edit_product(product):
    #Updates product information
    if request.method == "POST":
        with shelve.open("Databases/product") as f: 
            for i in f:
                if f[i]["name"] == product:
                    if request.form["name"] != "":
                        name = request.form["name"]
                    else:
                        name = f[i]["name"]
                    
                    if request.form["gender"] != "":
                        gender = request.form["gender"]
                    else:
                        gender = f[i]["gender"]
                    
                    if request.form["color"] != "":
                        color = request.form["color"]
                    else:
                        color = f[i]["color"]
                    
                    if request.form["categories"] != "":
                        categories = request.form["categories"]
                    else:
                        categories = f[i]["categories"]
                    
                    if request.form["RetailCost"] != "":
                        RetailCost = request.form["RetailCost"]
                    else:
                        RetailCost = f[i]["RetailCost"]
                    
                    if request.form["ListCost"] != "":
                        ListCost = request.form["ListCost"]
                    else:
                        ListCost = f[i]["ListCost"]

                    if request.form["s Quantity"] != "": 
                        s = request.form["s Quantity"]
                    else:
                        s = f[i]["stocks"]["S"]
                    
                    if request.form["m Quantity"] != "": 
                        m = request.form["m Quantity"]
                    else:
                        m = f[i]["stocks"]["M"]
                    
                    if request.form["l Quantity"] != "": 
                        l = request.form["l Quantity"]
                    else:
                        l = f[i]["stocks"]["L"]
                    
                    if request.form["xl Quantity"] != "": 
                        xl = request.form["xl Quantity"]
                    else:
                        xl = f[i]["stocks"]["XL"]

                    if request.form["xxl Quantity"] != "": 
                        xxl = request.form["xxl Quantity"]
                    else:
                        xxl = f[i]["stocks"]["XXL"] 

                    f[i] = {
                        "name" : name,
                        "image" : f"{name.replace(" ","_")}.jpeg",
                        "stocks" : {
                            "S" : s,
                            "M" : m,
                            "L" : l,
                            "XL" : xl,
                            "XXL" : xxl
                        },
                        "desc" : f[i]["desc"],
                        "points" : f[i]["points"],
                        "type" : f[i]["type"],
                        "gender":gender,
                        "categories" : categories,
                        "color":color,
                        "RetailCost" :RetailCost,
                        "ListCost" : ListCost,
                    }


            
    with shelve.open("Databases/product") as f:
        for i in f:
            if f[i]["name"] == product:
                name  = f[i]["name"]
                gender = f[i]["gender"]
                S = f[i]["stocks"]["S"]
                M = f[i]["stocks"]["M"]
                L  = f[i]["stocks"]["L"]
                Xl  = f[i]["stocks"]["XL"]
                Xxl  = f[i]["stocks"]["XXL"]
                total = int(S) + int(M) + int(L) + int(Xl) + int(Xxl)
                color = f[i]["color"]
                category = f[i]["categories"]
                retail  = f[i]["RetailCost"]
                list_cost  = f[i]["ListCost"]
                
    return render_template("admin_product_edit.html",name=name,gender=gender,S=S, M=M,L=L,Xl=Xl,Xxl=Xxl,total=total,color=color,category=category,retail=retail,list_cost=list_cost)

@app.route("/admin/userdata",methods=["POST","GET"])
def admin_user():
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            x = 0
            for i in f:
                if f[i]["username"] == request.form["submit"]:
                    del f[i]
                    return redirect("/admin/userdata")

    
                x += 1

    with shelve.open("Databases/user") as f:
        cards = ''
        for i in f:
            card1 = '''
                        <tr class="user-admin-tr">
                        <form method="post">
                            <td class="user-admin-td">
                                <div class="name">'''+ f[i]['username'] +'''</div>
                                <div class="email">'''+ f[i]['email'] +'''</div>
                            </td>
                            <td class="user-admin-td">'''+ f[i]['role'] +'''</td>
                            <td class="user-admin-td"><a href="
                            '''
            if f[i]["role"] == "user": 
                card2 = '''/admin/userdata/'''
            elif f[i]["role"] == "admin":
                card2 = "/admin/admindata/"
            card3 = f'''{f[i]["username"]}"><i class="fa fa-pencil" style="font-size:32px" id="update"></i></td>
                            <td class="user-admin-td">{f[i]['JoinDate']}</td>
                            <td class="user-admin-td"><button class="user-admin-button" type="submit" name="submit" value="{f[i]["username"]}"><i class="fa fa-trash-o" style="font-size:32px"></i></button></td>
                        </form></tr>
            '''
            card = card1 + card2 + card3
            cards += card
    return render_template("user-admin.html" ,data=cards)

@app.route("/admin/userdata/<string:Account>", methods=["POST", "GET"])
def admin_user_details(Account):
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            for id_ in f:
            
            
                if f[id_]["username"] == Account:
                    user_id = id_
           
            if request.form["username"] != "":
                username = request.form["username"]
            else: 
                username = f[user_id]["username"]
            
            if request.form["password"] != "":
                # validate password strength
                if re.match(r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)[a-zA-Z\d]{8,}$", password):
                    password = request.form["password"]
                else:
                    password = f[user_id]["password"]
                    flash(
                        "Invalid password. Password should contain at least one number, one uppercase and one lowercase letter, and be at least 8 characters long.")
            else: 
                password = f[user_id]["password"]

            if request.form["email"] != "":
                email = request.form["email"]
                # validate email format
                if re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email) and request.form["email"] !="":
                    email = request.form["email"]
                elif request.form["email"] == "":
                    flash("Please fill in Email Field")
                else:
                    flash("Invalid email address.")
            else: 
                email = f[user_id]["email"]
            
            if request.form["address"] != "":
                address = request.form["address"]
            else: 
                address = f[user_id]["address"]

            
            f[user_id] = {
                "username" : username,
                "password" : password,
                "role" : "user",
                "image" : username+".jpg",
                "email" : email,
                "address" : address,
                "Paynow" : f[user_id]["Paynow"],
                "Paypal" : f[user_id]["Paypal"],
                "JoinDate" : f[user_id]["JoinDate"],
                "point":f[user_id]["point"],
                "cart" : f[user_id]["cart"]
            }




    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == Account:
                username = f[i]["username"]
                email = f[i]["email"]
                address = f[i]["address"]
                password = f[i]["password"]
    return render_template("user-admin-edit.html", username=username, email=email, address=address, password=password)

@app.route("/admin/admindata/<string:Account>", methods=["POST","GET"])
def admin_edit_user(Account):
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == Account:
                    if request.form["username"] != "":
                        name = request.form["username"]
                        image = request.form["username"] + ".jpg"
                    else:
                        name = f[i]["username"]
                        image = f[i]["image"]
                    join_date = f[i]["JoinDate"]
                    role = "admin"
                    if request.form["email"] != "":
                        email = request.form["admin"]
                    else:
                        email = f[i]["email"]
                    if request.form["password"] != "":
                        password = request.form["password"]
                    else: 
                        password = f[i]["password"]


                    f[str(i)] = {
                        "username":name,
                        "password":password,
                        "JoinDate":join_date,
                        "image" : image,
                        "role":role,
                        "email":email
                    }
            return redirect(url_for("admin_user"))
        

    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == Account:
                username = f[i]["username"]
                email = f[i]["email"]
                password = f[i]["password"]
                join_date = f[i]["JoinDate"]

    return render_template("admin_admin_edit.html",username=username,email=email, password=password, account_date=join_date)



@app.route("/useredit/<string:Account>",methods=["POST","GET"])
def user_details(Account):
    if request.method == "POST":
        #Updates user information
        with shelve.open("Databases/user") as f:
            x = 0
            for i in f:
                if f[i]["username"] == Account:
                    if request.form["username"] != "":
                        username = request.form["username"]
                    else:
                        username = f[i]["username"]
                    if request.form["email"] != "":
                        email = request.form["email"]
                    else: 
                        email = f[i]["email"]
                    if request.form["password"] != "":
                        password = request.form["password"]
                    else:
                        password = f[i]["password"]
                if request.form["del"] != None:
                    #Deletes user information
                    del f[str(x)]
                f[i] = {
                    "username":username,
                    "password":password,
                    "role":"user",
                    "image":f"{username}.jpg",
                    "email":email,
                    "address":f[i]["address"],
                    "History":f[i]["History"],
                    "Paynow" : f[i]["Paynow"],
                    "Paypal" : f[i]["Paypal"],
                    "JoinDate":f[i]["JoinDate"],
                    "point":f[i]["point"],
                    "cart":f[i]["cart"]
                }
                x+= 1
    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == user.get_name():
                username = f[i]["username"]
                email = f[i]["email"]
                password = f[i]["password"]
                points = f[i]["point"]
    return render_template("settings-edit.html",username=username,email=email,password=password,points=points)

@app.route("/userdata/<string:Account>")
def user_data_view(Account):

    try:
        if Account == "":
            return render_template("keyerror.html")
        with shelve.open("Databases/user") as f:
                for i in f:
                    if f[i]["username"] == user.get_name():
                        
                        username = f[i]["username"]
                        join_date = f[i]["JoinDate"]
                        password = f[i]["password"]
                        image = f[i]["image"]
                        email = f[i]["email"]
                        address = f[i]["address"]
                        cart = f[i]["cart"]
                        points = f[i]["point"]
                        return render_template("settings.html",username = username,join_date = join_date,password=password, image = image,email=email,address=address,cart=cart,points=points)
    except: return render_template("keyerror.html")

    
@app.route("/user/rewards",methods=["GET","POST"])
def user_rewards():
    #Updates Cart
    #Retrieves reward items
    if request.method == "POST":
        reward_item = request.form['reward']
        with shelve.open("Databases/product") as f:
            for i in f:
                if f[i]["name"] == reward_item:
                    f[i]["stock"] -= 1
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["name"] == user.get_name():
                    f[i]["cart"] = f[i][cart].append(reward_item)
    
    with shelve.open("Databases/products") as f:
        data = ""
        for i in f:
            data = ""
            if f[i]["type"] == "reward":
                card = f'''<div class="reward">
                <img src="/static/Assets/{f[i]["image"]}" alt="{f[i]["name"]}">
                <div class="Info">
                    <header>{f[i]["name"]}</header>
                    <p>Description:
                        <br>
                        {f[i]["desc"]}
                    </p>
                    <p>{f[i]["points"]}</p>
                    <button name="redeem" value="{f[i]["name"]}">Redeem</button>
                    </div>
                    </div>'''
                data += card
    return render_template("reward.html",data=data)

@app.route("/admin/rewards",methods=["POST","GET"])
def admin_reward():
    if request.method == "POST":
        with shelve.open("Databases/product") as f:
            for i in f:
                #Deletes reward 
                if f[i]["name"] == request.form["submit"]:
                    del f[i]
                
        #Displays current reward
    cards = ""
    with shelve.open("Databases/product") as f:
        for i in f:
            card = ""
            if f[i]["type"] == "reward":
                card = f'''
                <tr> 
                    <td>{f[i]["name"]}</td>
                    <td>{f[i]["ListCost"]} pts</td>
                    <td>{f[i]["stock"]}</td>
                    <td>{f[i]["Status"]}</td>
                    <td><a href="/admin-rewards/{f[i]["name"]}">edit</a></td>
                 <td><form method="post">
                 <button type="submit" name="submit" value="{f[i]["name"]}"><i class="fa fa-trash-o" style="font-size:32px"></i></button>
                 </form>
                </td>
                </tr>
            '''
                cards += card
    return render_template("admin_rewards.html",cards=cards)

@app.route("/admin/reward-create")
def new_reward():
    if request.method == "POST":
        with shelve.open("Databases/product") as f:
            x = 0
            #Creates new reward
            for i in f:
                x += 1
            x+=1
            f[str(x)] = {
                "name" : request.form["name"],
                "image":request.form["image"],
                "stock":request.form["stocks"],
                "Desc":request.form["desc"],
                "type": request.form["type"],
                "ListCost" :request.form["points"],
                "Status" :request.form["status"]
            }
    return render_template("admin_reward_create.html")

@app.route("/admin/home")
def admin_home():
    
    #Transaction, User, Product
    wb = openpyxl.Workbook()    
    ws1 = "Transaction"
    ws2 = "User"
    ws3 = "product"

    wb.create_sheet(ws1)
    wb.create_sheet(ws2)
    wb.create_sheet(ws3)

    with shelve.open("Databases/transaction") as f:
        #Titles
        x = 1
        wb[ws1]["A1"] = "id"
        wb[ws1]["B1"] = "Date"
        wb[ws1]["C1"] = "Product"
        wb[ws1]["D1"] = "Username"
        wb[ws1]["E1"] = "Status"
        wb[ws1]["F1"] = "Price"
        for i in f:
            x += 1
            wb[ws1]["A"+str(x)] = f[i]["id"]
            wb[ws1]["B"+str(x)] = f[i]["Date"]
            wb[ws1]["C"+str(x)] = f[i]["Product"]
            wb[ws1]["D"+str(x)] = f[i]["Cust"]
            wb[ws1]["E"+str(x)] = f[i]["Status"]
            wb[ws1]["F"+str(x)] = f[i]["Price"]
    with shelve.open("Databases/user") as f:
        x = 1
        wb[ws2]["A1"] = "Username"
        wb[ws2]["B1"] = "password"
        wb[ws2]["C1"] = "role"
        wb[ws2]["D1"] = "email"
        wb[ws2]["E1"] = "address"
        wb[ws2]["F1"] = "JoinDate"
        wb[ws2]["G1"] = "point"
        wb[ws2]["H1"] = "cart"
        for i in f:
            x += 1
            if f[i]["role"] == "admin":
                wb[ws2]["A"+str(x)] = f[i]["username"]
                wb[ws2]["B"+str(x)] = f[i]["password"]
                wb[ws2]["C"+str(x)] = f[i]["role"]
                wb[ws2]["D"+str(x)] = f[i]["email"]
                wb[ws2]["F"+str(x)] = f[i]["JoinDate"]
            
            elif f[i]["role"] == "user":
                wb[ws2]["A"+str(x)] = f[i]["username"]
                wb[ws2]["B"+str(x)] = f[i]["password"]
                wb[ws2]["C"+str(x)] = f[i]["role"]
                wb[ws2]["D"+str(x)] = f[i]["email"]
                wb[ws2]["E"+str(x)] = f[i]["address"]
                wb[ws2]["F"+str(x)] = f[i]["JoinDate"]
                wb[ws2]["G"+str(x)] = str(f[i]["point"])
                wb[ws2]["H"+str(x)] = str(f[i]["cart"])
    with shelve.open("Databases/product") as f:
        x = 1
        wb[ws3]["A1"] = "name"
        wb[ws3]["B1"] = "Size S"
        wb[ws3]["C1"] = "Size M"
        wb[ws3]["D1"] = "Size L"
        wb[ws3]["E1"] = "Size XL"
        wb[ws3]["F1"] = "Size XXL"
        wb[ws3]["G1"] = "desc"
        wb[ws3]["H1"] = "points"
        wb[ws3]["I1"] = "type"
        wb[ws3]["J1"] = "gender"
        wb[ws3]["K1"] = "Categories"
        wb[ws3]["L1"] = "color"
        wb[ws3]["M1"] = "RetailCost"
        wb[ws3]["N1"] = "ListCost"
        for i in f:
            x += 1
            if f[i]["type"] == "product":
                wb[ws3]["A"+str(x)] = f[i]["name"]
                wb[ws3]["B"+str(x)] = f[i]["stocks"]["S"]
                wb[ws3]["C"+str(x)] = f[i]["stocks"]["M"]
                wb[ws3]["D"+str(x)] = f[i]["stocks"]["L"]
                wb[ws3]["E"+str(x)] = f[i]["stocks"]["XL"]
                wb[ws3]["F"+str(x)] = f[i]["stocks"]["XXL"]
                wb[ws3]["G"+str(x)] = f[i]["desc"]
                wb[ws3]["H"+str(x)] = f[i]["points"]
                wb[ws3]["I"+str(x)] = f[i]["type"]
                wb[ws3]["J"+str(x)] = f[i]["gender"]
                wb[ws3]["K"+str(x)] = f[i]["categories"]
                wb[ws3]["L"+str(x)] = f[i]["color"]
                wb[ws3]["M"+str(x)] = f[i]["RetailCost"]
                wb[ws3]["N"+str(x)] = f[i]["ListCost"]
            elif f[i]["type"] == "reward":
                wb[ws3]["A"+str(x)] = f[i]["name"]
                wb[ws3]["G"+str(x)] = f[i]["Desc"]
                wb[ws3]["H"+str(x)] = f[i]["ListCost"]
                wb[ws3]["I"+str(x)] = f[i]["type"]
    wb.save("static/report/Reville.xlsx")
    with shelve.open("Databases/transaction") as f:
        list1 = [0,0,0,0,0,0,0,0,0,0,0,0]
        for i in f:
            month = f[i]["Date"].split("-")[1]
            list1[int(month)] += 1

        list2 = str(" ".join(str(i) for i in list1))
            
    return render_template("admin_home.html",username=user.get_name(),list1=list2)

    
@app.route('/admin/transaction', methods=["POST","GET"])
def admin_transaction():
    if request.method == "POST":
        with shelve.open("Databases/transaction") as f:
            for i in f:
                try:
                    print(request.form["delete"],f[i]["id"])
                    if request.form["delete"] == str(f[i]["id"]):
                        print("deleted")
                        del f[i]
                        return redirect(url_for("admin_transaction"))
                except: 
                    print(f[i])
                    if request.form["edit"] == str(f[i]["id"]):

                        print("edited")
                        status = "Delivered"
                        f[i] = {
                                "id" : i,
                                "Date" : f[i]["Date"],
                                "Product" : f[i]["Product"],
                                "Cust" : f[i]["Cust"],
                                "Status" : status,
                                "Price" : f[i]["Price"]
                            }
                        return redirect(url_for("admin_transaction"))
                

    with shelve.open('Databases/transaction') as f:
        rows = ""
        x = 0
        for i in f:
            row = f'''
            <tr> 
                <td>{f[i]["id"]}</td>
                <td>{f[i]["Date"]}</td>
                <td>{f[i]["Product"].replace("_"," ")}</td>
                <td>{f[i]["Cust"]}</td>
                <td>{f[i]["Status"]}</td>
                <td>{f[i]["Price"]}</td>
                <td><form method="post"><button type="submit" name="edit" value="{f[i]['id']}">Edit</button></form></td>
                <td><form method="post"><button type="submit" name="delete" value="{f[i]['id']}">Delete</button></form></td>
            </tr>'''
            rows += row
            x += 1
    return render_template("admin_transaction.html",data=rows,no=x)


@app.errorhandler(500)
def internal_error(e):
    return render_template("keyerror.html"), 500

@app.errorhandler(404)
def not_found(e):
    return render_template("keyerror.html"),404


@app.route("/address/<string:Account>")
def addresses(Account):
    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == Account:
                address = f[i]["address"]
                points = f[i]["point"]
    return render_template("delivery.html",username=Account, address=address,points=points)

@app.route("/address/edit/<string:Account>", methods=["POST","GET"])
def address_edit(Account):
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == Account:
                    if request.form["delivery-address"] != "":
                        f[i] = {
                        "username":f[i]["username"],
                        "password":f[i]["password"],
                        "role":"user",
                        "image":f[i]["image"],
                        "email":f[i]["email"],
                        "address":request.form["delivery-address"],
                        "Paynow" : f[i]["Paynow"],
                        "Paypal" : f[i]["Paypal"],
                        "JoinDate":f[i]["JoinDate"],
                        "point":f[i]["point"],
                        "cart": f[i]["cart"]
                        }
                    return redirect(url_for(f"addresses({Account})"))
    with shelve.open("Databases/user") as f:
        for i in f:
            if f[i]["username"] == Account:
                address = f[i]["address"]
                points = f[i]["point"]
    return render_template("delivery.html",username=Account, address=address,points=points)

@app.route("/paynow/<string:Account>")
def paynow(Account):
    try:
        with shelve.open("Databases/user") as f:
            try:
                for i in f:
                    if f[i]["username"] == user.get_name():
                        try: 
                            name = f[i]["Paynow"]["name"]
                        except : 
                            name = "No registered Name"
                        try:
                            number = f[i]["Paynow"]["number"]
                        except:
                            number = "No registered Account Number"
                        try:
                            cvv = f[i]["Paynow"]["cvv"]
                        except:
                            cvv = "No registered cvv"
                        points = f[i]["point"]
                        break
                        
            except:
                print("hellos")
                return render_template("keyerror.html")
                
        return render_template("payment-paynow.html",name=name,number=number,cvv=cvv,username=Account,points=points)
    except:
        return render_template("keyerror.html")

@app.route("/paypal/<string:Account>")
def paypal(Account):    
    try:
        with shelve.open("Databases/user") as f:
            try:
                for i in f:
                    if f[i]["username"] == user.get_name():
                        try: 
                            name = f[i]["Paypal"]["email"]
                        except : 
                            name = "No registered email"
                        try:
                            number = f[i]["Paypal"]["password"]
                        except:
                            number = "No registered Account Number"
                        points = f[i]["point"]
                        break
            except:
                return render_template("keyerror.html")
                
        return render_template("payment-paypal.html",email=name,password=number,username=Account,points=points)
    except:
        return render_template("keyerror.html")

@app.route("/payment_edit/paypal/<string:Account>",methods=["POST","GET"])
def paypal_edit(Account):
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == user.get_name():
                    if request.form["paypal-email"] != "":
                        email = request.form["paypal-email"]
                    else:
                        email = f[i]["Paypal"]["email"]
                                
                    if request.form["paypal-password"] != "":
                        password = request.form["paypal-password"]
                    else:
                        password = f[i]["Paypal"]["password"]
                    f[i] = {
                        "username":f[i]["username"],
                        "password":f[i]["password"],
                        "role":"user",
                        "image":f[i]["image"],
                        "email":f[i]["email"],
                        "address":f[i]["address"],
                        "Paynow" : f[i]["Paynow"],
                        "Paypal" : {"email":email,"password":password},
                        "JoinDate":f[i]["JoinDate"],
                        "point":f[i]["point"],
                        "cart": f[i]["cart"]
                        }
                    return redirect(url_for("paypal"))
    try: 
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == user.get_name():
                    email = f[i]["Paypal"]["email"]
                    password = f[i]["Paypal"]["password"]
                    points = f[i]["point"]
                    break
        return render_template("payment-paypal-edit.html",email=email,password=password,username=Account,points=points)
    except: return render_template("keyerror.html")



@app.route("/payment_edit/paynow/<string:Account>",methods=["POST","GET"])
def paynow_edit(Account):
    if request.method == "POST":
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == user.get_name():
                    if request.form["paynow-name"] != "":
                        name = request.form["paynow-name"]
                    else:
                        name = f[i]["Paynow"]["name"]
                                
                    if request.form["paynow-cvv"] != "":
                        cvv = request.form["paynow-cvv"]
                    else:
                        cvv = f[i]["Paynow"]["cvv"]
                    
                    if request.form["paynow-account_num"] != "":
                        account_num = request.form["paynow-account_num"]
                    else:
                        account_num = f[i]["Paynow"]["account_num"]
                    f[i] = {
                        "username":f[i]["username"],
                        "password":f[i]["password"],
                        "role":"user",
                        "image":f[i]["image"],
                        "email":f[i]["email"],
                        "address":f[i]["address"],
                        "Paynow" : {"name":name,"number":account_num,"cvv":cvv},
                        "Paypal" : f[i]["Paypal"],
                        "JoinDate":f[i]["JoinDate"],
                        "point":f[i]["point"],
                        "cart": f[i]["cart"]
                        }
                    return redirect(url_for("paynow"))
    try: 
        with shelve.open("Databases/user") as f:
            for i in f:
                if f[i]["username"] == user.get_name():
                    name = f[i]["Paynow"]["name"]
                    account_num = f[i]["Paynow"]["number"]
                    cvv = f[i]["Paynow"]["cvv"]
                    points = f[i]["point"]
                    break
        return render_template("payment-paynow-edit.html",name = name,account_num = account_num , cvv=cvv, username=Account,points=points)
    except: return render_template("keyerror.html")

@app.route("/transaction/<string:Account>")
def user_transaction(Account):
    cards = ""
    with shelve.open("Databases/transaction") as f:
        for i in f:
            if f[i]["Cust"] == user.get_name():
                with shelve.open("Databases/product") as g:
                    for ii in g:
                        print(g[ii]["name"], f[i]["Product"].replace("_"," "))
                        if g[ii]["name"] == str(f[i]["Product"]).replace("_"," "):
                            name = g[ii]["name"]
                            _type = g[ii]["type"]
                            price = f[i]["Price"]
                            date = f[i]["Date"]
                            try:
                                points = g[ii]["points"]
                            except:
                                points = "Nil"
                            card = f'''
                                <tr>
                                    <td>{name }</td><input type="hidden" name="name">
                                    <td>{ _type }</td><input type="hidden" name="type">
                                    <td>{ price }</td><input type="hidden" name="price">
                                    <td>{ date }</td><input type="hidden" name="date">
                                    <td>
                                        {f[i]["Status"]}
                                        <input type="hidden" name="status" value="Active">
                                    </td>
                                </tr>
                                    '''
                            cards+= card
                            break
    return render_template("transaction-history.html",username=Account, data=cards,points=points )


if __name__ == "__main__":
    initate()
    app.config.update(
    SECRET_KEY='71d9d79dc6879448be43c10b5289436ae90bd76c56d170b824')
    app.run()