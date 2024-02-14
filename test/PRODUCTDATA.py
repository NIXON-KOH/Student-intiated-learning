import shelve
import openpyxl
#DONT EDIT THE FILE OR I WILL EDIT YOU OUT OF EARTH :D

import os
os.system('cls' if os.name == 'nt' else 'clear')

wb = openpyxl.load_workbook("Product_info.xlsx")
ws = wb.active 
Products = []
for i in range(1,ws.max_row+1):
    col_ = []
    for n in range(1,ws.max_column+1):
        cell = ws.cell(row=i,column=n)
        col_.append(cell.value)
    Products.append(col_)

Products.pop(0)
Products.pop(-1)
# id_, name,Image, Stocks, Desc, Type, Status,categories,RetailCost,gender,ListCost,reviews
x = 0
with shelve.open("Databases/product") as f:
    for i in Products:
        x += 1
        f[str(x)] = {
            "name" : i[0],
            "image" : f"{i[0]}.jpeg",
            "stocks" : {
                "S" : i[5],
                "M" : i[6],
                "L" : i[7],
                "XL" : i[8],
                "XXL" : i[9]
            },
            "desc" : i[-1],
            "points" : 100,
            "type" : "product",
            "gender":i[1],
            "categories" : i[2],
            "color":i[3],
            "RetailCost" : 50,
            "ListCost" : 100,
        }
        print(f"{x} Complete!")
    #id_, name,Image, Stocks, Desc, Type, points
    x += 1
    f[str(x)] = {
        "name" : "Resuable Lunchbox",
        "image" : "defualt-image.jpg",
        "stock" : 4,
        "Desc" : 'PREMIUM Bamboo Lunch Box. This lunchbox uses environmentally friendly plant based materials that can be reused again & again. Bamboo is a fast growing plant & our bamboo items are from managed bamboo plantations.',
        "type" : "reward",
        "ListCost" : 100,
        "Status" : "Active"
        }
    print(f"{x} Complete!")
    x += 1
    f[str(x)] = {
        "name" : "Water Bottle",
        "image" : "defualt-image.jpg",
        "stock" : 10,
        "Desc" : 'PREMIUM Stainless Steel Water Bottle. This water bottle is environmentally friendly and can be reused again & again.',
        "type" : "reward",
        "ListCost" : 150,
        "Status" : "Active"
        }
    print(f"{x} Complete!")


    x += 1
    f[str(x)] = {
        "name" : "Tote Bag",
        "image" : "defualt-image.jpg",
        "stock" : 10,
        "Desc" : 'Reusable tote bag so that we can help save the environment by decreasing our use of plastic bags!',
        "type" : "reward",
        "ListCost" : 300,
        "Status" : "Inactive"
        }
    print(f"{x} Complete!")


    x += 1
    f[str(x)] = {
        "name" : "Cutlery Set",
        "image" : "defualt-image.jpg",
        "stock" : 10,
        "Desc" : 'Cutlery set so that we can cut down on the use of disposable cutlery!',
        "type" : "reward",
        "ListCost" : 500,
        
        "Status" : "Inactive"   
        }
    print(f"{x} Complete!")


print("ALL COMPLETE!")